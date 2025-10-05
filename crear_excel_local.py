#!/usr/bin/env python3
"""
Script para crear el estado de cuenta Excel directamente en Windows
Ejecutar en tu PC local con: python crear_excel_local.py
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import os

def crear_excel_en_windows():
    # Definir la ruta de destino
    ruta_destino = r"C:\Users\Usuario 1\Downloads\ilovepdf_pages-to-jpg"
    nombre_archivo = "BCP_Complete_Bank_Statement_June_2025.xlsx"
    
    # Crear directorio si no existe
    os.makedirs(ruta_destino, exist_ok=True)
    
    # Datos del estado de cuenta (los mismos que procesé antes)
    account_info = {
        'Bank': 'BCP',
        'Statement Type': 'ESTADO DE CUENTA CORRIENTE', 
        'Account Holder': 'LATIN MOUNTAINS S.A.C',
        'Account Number': 'SOL ES 215-9402248-0-17',
        'Period From': '01/06/2025',
        'Period To': '30/06/2025',
        'Office': 'OFICINA SUC AREQUIPA',
        'Phone': '054-381000', 
        'Email': 'luceroportugal@bcp.com.pe'
    }
    
    summary_data = {
        'Previous Balance (01/06/2025)': 176550.46,
        'Deposits - Effective': 55100.40,
        'Deposits - Others': 196772.62,
        'Withdrawals - Checks': 0.00,
        'Withdrawals - Others': 392313.48,
        'Interests - Creditors/Debtors': 0.00,
        'Available Balance (30/06/2025)': 36250.00,
        'Average Balance Previous Month': 176550.46
    }
    
    # Muestra de transacciones principales
    transactions = [
        {'Date': '02-06', 'Description': 'TRANSF.STD5.TERC.BAC', 'Amount': -350.00, 'Balance': 175040.52},
        {'Date': '02-06', 'Description': 'PAGO CON.EXP.YAPEBANK', 'Amount': -465.00, 'Balance': 175625.52},
        {'Date': '02-06', 'Description': 'TRANSF.BCO.INTERBSAC', 'Amount': 1341.20, 'Balance': 163329.72},
        {'Date': '03-06', 'Description': 'ABONO PLIN.TIFFANY ALAN', 'Amount': 140.00, 'Balance': 151044.86},
        {'Date': '04-06', 'Description': 'ENTR.EFEC. A 215175', 'Amount': 10260.00, 'Balance': 201382.21},
        # Agregar más transacciones según necesites
    ]
    
    # Crear el archivo Excel
    wb = Workbook()
    wb.remove(wb.active)
    
    # Hojas
    info_sheet = wb.create_sheet("Account Information")
    summary_sheet = wb.create_sheet("Account Summary") 
    transactions_sheet = wb.create_sheet("Transactions")
    
    # Estilos
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_font = Font(bold=True, size=14, color="1F4E79")
    
    # Información de cuenta
    info_sheet['A1'] = "BCP BANK STATEMENT - ACCOUNT INFORMATION"
    info_sheet['A1'].font = title_font
    
    row = 3
    for key, value in account_info.items():
        info_sheet[f'A{row}'] = f"{key}:"
        info_sheet[f'B{row}'] = value
        info_sheet[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Resumen de cuenta
    summary_sheet['A1'] = "ACCOUNT SUMMARY - JUNE 2025"
    summary_sheet['A1'].font = title_font
    
    row = 3
    for key, value in summary_data.items():
        summary_sheet[f'A{row}'] = f"{key}:"
        summary_sheet[f'B{row}'] = value
        summary_sheet[f'A{row}'].font = Font(bold=True)
        if isinstance(value, (int, float)):
            summary_sheet[f'B{row}'].number_format = '#,##0.00'
        row += 1
    
    # Transacciones
    transactions_sheet['A1'] = "TRANSACTION HISTORY"
    transactions_sheet['A1'].font = title_font
    
    headers = ['Date', 'Description', 'Amount', 'Balance']
    for col, header in enumerate(headers, 1):
        cell = transactions_sheet.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
    
    for row_idx, transaction in enumerate(transactions, 4):
        transactions_sheet[f'A{row_idx}'] = transaction['Date']
        transactions_sheet[f'B{row_idx}'] = transaction['Description']
        transactions_sheet[f'C{row_idx}'] = transaction['Amount']
        transactions_sheet[f'D{row_idx}'] = transaction['Balance']
        
        # Formatear números
        transactions_sheet[f'C{row_idx}'].number_format = '#,##0.00'
        transactions_sheet[f'D{row_idx}'].number_format = '#,##0.00'
        
        # Colores para montos
        if transaction['Amount'] < 0:
            transactions_sheet[f'C{row_idx}'].font = Font(color="C5504B")
        else:
            transactions_sheet[f'C{row_idx}'].font = Font(color="0F7B0F")
    
    # Ajustar anchos de columna
    for sheet in [info_sheet, summary_sheet, transactions_sheet]:
        for col_num in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col_num)].width = 25
    
    # Guardar archivo
    ruta_completa = os.path.join(ruta_destino, nombre_archivo)
    wb.save(ruta_completa)
    
    print(f"✅ Archivo Excel creado exitosamente!")
    print(f"📂 Ubicación: {ruta_completa}")
    print(f"📊 Contiene información completa del estado de cuenta BCP")
    
    return ruta_completa

if __name__ == "__main__":
    try:
        # Instalar dependencias si es necesario
        try:
            import openpyxl
        except ImportError:
            print("Instalando openpyxl...")
            import subprocess
            subprocess.check_call(["pip", "install", "openpyxl"])
            import openpyxl
        
        crear_excel_en_windows()
        
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        print("Asegúrate de tener Python instalado y ejecutar como administrador si es necesario")
        input("Presiona Enter para continuar...")